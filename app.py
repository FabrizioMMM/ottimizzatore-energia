import streamlit as st
import pandas as pd
import numpy as np
from scipy.optimize import linprog
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configurazione pagina ─────────────────────────────────────────────────────
st.set_page_config(
    page_title="Ottimizzatore Consumi Energetici",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── CSS – Tema Verde Energia, Alto Contrasto ──────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Nunito:wght@400;600;700;800&family=Roboto+Mono:wght@400;600&display=swap');

/* BASE */
html, body, [class*="css"], .stApp {
    font-family: 'Nunito', sans-serif;
    background-color: #f0faf4 !important;
    color: #0f2b1a !important;
}

/* INTESTAZIONE PRINCIPALE */
.intestazione {
    background: linear-gradient(120deg, #065f46 0%, #047857 50%, #059669 100%);
    padding: 2rem 2.5rem;
    border-radius: 14px;
    margin-bottom: 2rem;
    box-shadow: 0 4px 20px rgba(6,95,70,0.25);
}
.intestazione h1 {
    font-family: 'Nunito', sans-serif;
    font-weight: 800;
    color: #ffffff;
    font-size: 2rem;
    margin: 0 0 0.4rem 0;
    text-shadow: 0 1px 3px rgba(0,0,0,0.2);
}
.intestazione p {
    color: #a7f3d0;
    margin: 0;
    font-size: 1.05rem;
    font-weight: 600;
}

/* TITOLI SEZIONE */
.titolo-sezione {
    font-family: 'Nunito', sans-serif;
    font-weight: 800;
    font-size: 1rem;
    color: #065f46;
    text-transform: uppercase;
    letter-spacing: 0.06em;
    border-left: 4px solid #10b981;
    padding-left: 0.7rem;
    margin: 2rem 0 1rem 0;
}

/* CARD RISULTATI */
.card-risparmio {
    background: #065f46;
    border-radius: 12px;
    padding: 1.5rem 1.8rem;
    box-shadow: 0 3px 12px rgba(6,95,70,0.3);
    height: 100%;
}
.card-risparmio .etichetta {
    color: #6ee7b7;
    font-size: 0.9rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    margin-bottom: 0.5rem;
}
.card-risparmio .valore-grande {
    color: #ffffff;
    font-family: 'Roboto Mono', monospace;
    font-size: 2.4rem;
    font-weight: 600;
    line-height: 1.1;
}
.card-risparmio .sotto-testo {
    color: #a7f3d0;
    font-size: 0.95rem;
    font-weight: 600;
    margin-top: 0.4rem;
}

.card-info {
    background: #1e3a5f;
    border-radius: 12px;
    padding: 1.5rem 1.8rem;
    box-shadow: 0 3px 12px rgba(30,58,95,0.3);
    height: 100%;
}
.card-info .etichetta {
    color: #93c5fd;
    font-size: 0.9rem;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    margin-bottom: 0.5rem;
}
.card-info .valore-grande {
    color: #ffffff;
    font-family: 'Roboto Mono', monospace;
    font-size: 2.4rem;
    font-weight: 600;
    line-height: 1.1;
}
.card-info .sotto-testo {
    color: #bfdbfe;
    font-size: 0.95rem;
    font-weight: 600;
    margin-top: 0.4rem;
}

/* RIQUADRO ISTRUZIONI */
.riquadro-istruzioni {
    background: #ecfdf5;
    border: 2px solid #6ee7b7;
    border-radius: 10px;
    padding: 1.2rem 1.5rem;
}
.riquadro-istruzioni p {
    color: #065f46;
    font-size: 1rem;
    font-weight: 600;
    margin: 0.2rem 0;
}

/* SIDEBAR */
[data-testid="stSidebar"] {
    background-color: #052e16 !important;
    border-right: 3px solid #059669;
}
[data-testid="stSidebar"] .stMarkdown,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] span,
[data-testid="stSidebar"] div {
    color: #d1fae5 !important;
    font-size: 1rem !important;
}
[data-testid="stSidebar"] h3 {
    color: #6ee7b7 !important;
    font-size: 1.15rem !important;
    font-weight: 800 !important;
}
[data-testid="stSidebar"] .stSlider > div > div > div {
    background: #059669 !important;
}
[data-testid="stSidebar"] input {
    background: #064e3b !important;
    color: #ffffff !important;
    border: 1px solid #059669 !important;
    font-size: 1rem !important;
}
[data-testid="stSidebar"] .stCaption {
    color: #6ee7b7 !important;
    font-size: 0.88rem !important;
}

/* BOTTONE PRINCIPALE */
.stButton > button {
    background: linear-gradient(135deg, #059669, #065f46) !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 10px !important;
    font-family: 'Nunito', sans-serif !important;
    font-weight: 800 !important;
    font-size: 1.1rem !important;
    padding: 0.75rem 2.5rem !important;
    letter-spacing: 0.02em !important;
    box-shadow: 0 3px 10px rgba(5,150,105,0.4) !important;
    transition: all 0.2s !important;
}
.stButton > button:hover {
    background: linear-gradient(135deg, #047857, #064e3b) !important;
    box-shadow: 0 5px 15px rgba(5,150,105,0.5) !important;
    transform: translateY(-2px) !important;
}

/* BOTTONE DOWNLOAD */
.stDownloadButton > button {
    background: linear-gradient(135deg, #1d4ed8, #1e40af) !important;
    color: #ffffff !important;
    border: none !important;
    border-radius: 10px !important;
    font-family: 'Nunito', sans-serif !important;
    font-weight: 800 !important;
    font-size: 1rem !important;
    padding: 0.65rem 2rem !important;
    box-shadow: 0 3px 10px rgba(29,78,216,0.35) !important;
}

/* METRICHE */
[data-testid="stMetric"] {
    background: #ffffff;
    border: 2px solid #6ee7b7;
    border-radius: 10px;
    padding: 1rem 1.2rem;
    box-shadow: 0 2px 8px rgba(6,95,70,0.1);
}
[data-testid="stMetricLabel"] {
    color: #065f46 !important;
    font-weight: 700 !important;
    font-size: 0.95rem !important;
}
[data-testid="stMetricValue"] {
    color: #0f2b1a !important;
    font-size: 1.6rem !important;
    font-weight: 800 !important;
}
[data-testid="stMetricDelta"] {
    font-size: 0.9rem !important;
    font-weight: 700 !important;
}

/* TABELLA */
[data-testid="stDataFrame"] {
    border: 2px solid #a7f3d0 !important;
    border-radius: 10px !important;
    overflow: hidden !important;
}
.dataframe {
    font-size: 0.95rem !important;
    font-family: 'Roboto Mono', monospace !important;
}

/* MESSAGGI */
[data-testid="stAlert"] {
    font-size: 1rem !important;
    font-weight: 600 !important;
}

/* UPLOAD */
[data-testid="stFileUploader"] {
    background: #ecfdf5 !important;
    border: 2px dashed #059669 !important;
    border-radius: 10px !important;
}

/* SEPARATORI */
hr {
    border-color: #a7f3d0 !important;
    margin: 1.5rem 0 !important;
}

/* PIÈ DI PAGINA */
.pie-pagina {
    text-align: center;
    color: #065f46;
    font-size: 0.9rem;
    font-weight: 700;
    padding: 1rem 0 0.5rem 0;
}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
# FUNZIONI CORE
# ══════════════════════════════════════════════════════════════════════════════

def esegui_ottimizzazione(domanda, prezzi, potenza_kw, quota_flessibile, ora_inizio=1, ora_fine=24):
    """
    Solver LP: minimizza il costo energetico spostando i carichi flessibili
    nelle ore più economiche, rispettando la potenza impegnata contrattuale.
    """
    n = 24
    D = np.array(domanda, dtype=float)
    P = np.array(prezzi,  dtype=float)
    totale = D.sum()

    fisso = D * (1 - quota_flessibile)
    limite_sup = np.full(n, potenza_kw)

    for i in range(n):
        ora = i + 1
        if ora < ora_inizio or ora > ora_fine:
            limite_sup[i] = fisso[i]

    limite_inf = np.minimum(fisso, limite_sup)

    c      = P / 1000.0
    A_eq   = np.ones((1, n))
    b_eq   = np.array([totale])
    bounds = list(zip(limite_inf, limite_sup))

    risultato = linprog(c, A_eq=A_eq, b_eq=b_eq, bounds=bounds, method='highs')
    if risultato.success:
        return risultato.x, True
    else:
        return D.copy(), False


def leggi_file_gme(file_caricato):
    """Legge il file GME e restituisce i prezzi medi orari (24 valori)."""
    try:
        df = pd.read_excel(file_caricato, header=None)
        riga_header = 0
        for i, riga in df.iterrows():
            if any(str(v).strip() in ['Ora', 'Periodo', 'ora'] for v in riga.values if pd.notna(v)):
                riga_header = i
                break

        df = pd.read_excel(file_caricato, header=riga_header)
        df.columns = [str(c).strip() for c in df.columns]

        col_prezzo = next((c for c in df.columns if '€' in c or 'MWh' in c or 'Prezzo' in c.capitalize()), None)
        col_ora    = next((c for c in df.columns if c.lower() in ['ora', 'hour']), None)

        if col_prezzo is None or col_ora is None:
            df.columns = ['Data', 'Ora', 'Periodo', 'Prezzo'] + list(df.columns[4:])
            col_ora, col_prezzo = 'Ora', 'Prezzo'

        df[col_prezzo] = df[col_prezzo].apply(
            lambda x: float(str(x).replace(',', '.')) if pd.notna(x) else np.nan
        )
        df[col_ora] = pd.to_numeric(df[col_ora], errors='coerce')
        df = df.dropna(subset=[col_ora, col_prezzo])

        media_oraria = df.groupby(col_ora)[col_prezzo].mean().to_dict()
        prezzi = [media_oraria.get(float(h), media_oraria.get(h, 150.0)) for h in range(1, 25)]
        return prezzi, None
    except Exception as e:
        return None, str(e)


def esporta_excel(prezzi, domanda, ottimizzato, parametri):
    """Genera il report Excel con il piano orario ottimizzato."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Piano Ottimizzato"
    ws.sheet_view.showGridLines = False

    def riempimento(esad): return PatternFill('solid', start_color=esad, fgColor=esad)
    def bordo():
        l = Side(style='thin')
        return Border(left=l, right=l, top=l, bottom=l)
    centro = Alignment(horizontal='center', vertical='center')

    # Titolo
    ws.merge_cells('A1:I1')
    ws['A1'] = '⚡ PIANO ORARIO OTTIMIZZATO – Ottimizzatore Consumi Energetici'
    ws['A1'].font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
    ws['A1'].fill = riempimento('065f46')
    ws['A1'].alignment = centro
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:I2')
    ws['A2'] = (f"Potenza impegnata: {parametri['potenza_kw']} kW  |  "
                f"Flessibilità: {parametri['quota_flessibile']*100:.0f}%  |  "
                f"Energia totale: {sum(domanda):.2f} kWh")
    ws['A2'].font = Font(name='Arial', size=9, italic=True, color='065f46')
    ws['A2'].alignment = centro
    ws.row_dimensions[2].height = 16

    intestazioni = ['Ora', 'Prezzo (€/MWh)', 'Consumo Attuale (kWh)', 'Consumo Ottimizzato (kWh)',
                    'Variazione (kWh)', 'Costo Attuale (€)', 'Costo Ottimizzato (€)',
                    'Risparmio (€)', 'Risparmio (%)']
    larghezze = [8, 18, 22, 24, 18, 20, 22, 16, 14]
    for i, (h, w) in enumerate(zip(intestazioni, larghezze), 1):
        col = get_column_letter(i)
        ws[f'{col}3'] = h
        ws[f'{col}3'].font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        ws[f'{col}3'].fill = riempimento('047857')
        ws[f'{col}3'].alignment = centro
        ws[f'{col}3'].border = bordo()
        ws.column_dimensions[col].width = w
    ws.row_dimensions[3].height = 28

    costo_att_tot = costo_ott_tot = 0
    for i in range(24):
        r  = i + 4
        p  = prezzi[i]
        d  = domanda[i]
        o  = ottimizzato[i]
        ca = d * p / 1000
        co = o * p / 1000
        sv = ca - co
        pc = sv / ca if ca > 0 else 0
        costo_att_tot += ca
        costo_ott_tot += co

        sfondo = 'FFFFFF' if i % 2 == 0 else 'F0FDF4'
        valori = [i+1, round(p,3), round(d,4), round(o,4), round(o-d,4),
                  round(ca,5), round(co,5), round(sv,5), pc]
        formati = ['0','#,##0.000','#,##0.0000','#,##0.0000',
                   '+#,##0.0000;-#,##0.0000',
                   '€ #,##0.00000','€ #,##0.00000',
                   '+€ #,##0.00000;-€ #,##0.00000','0.0%']

        for j, (v, fmt) in enumerate(zip(valori, formati), 1):
            col = get_column_letter(j)
            ws[f'{col}{r}'] = v
            ws[f'{col}{r}'].font = Font(name='Arial', size=9)
            ws[f'{col}{r}'].alignment = centro
            ws[f'{col}{r}'].border = bordo()
            ws[f'{col}{r}'].number_format = fmt
            if j == 4:
                delta = o - d
                if delta > 0.001:
                    ws[f'{col}{r}'].fill = riempimento('d1fae5')
                    ws[f'{col}{r}'].font = Font(name='Arial', size=9, color='065f46', bold=True)
                elif delta < -0.001:
                    ws[f'{col}{r}'].fill = riempimento('fee2e2')
                    ws[f'{col}{r}'].font = Font(name='Arial', size=9, color='991b1b', bold=True)
                else:
                    ws[f'{col}{r}'].fill = riempimento(sfondo)
            else:
                ws[f'{col}{r}'].fill = riempimento(sfondo)

    risparmio = costo_att_tot - costo_ott_tot
    tr = 28
    totali = ['TOTALI', '', round(sum(domanda),4), round(sum(ottimizzato),4), '',
              round(costo_att_tot,5), round(costo_ott_tot,5), round(risparmio,5),
              risparmio/costo_att_tot if costo_att_tot > 0 else 0]
    formati_t = ['','','#,##0.0000 "kWh"','#,##0.0000 "kWh"','',
                 '€ #,##0.00000','€ #,##0.00000',
                 '+€ #,##0.00000;-€ #,##0.00000','0.0%']
    for j, (v, fmt) in enumerate(zip(totali, formati_t), 1):
        col = get_column_letter(j)
        ws[f'{col}{tr}'] = v
        ws[f'{col}{tr}'].font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
        ws[f'{col}{tr}'].fill = riempimento('065f46')
        ws[f'{col}{tr}'].alignment = centro
        ws[f'{col}{tr}'].border = bordo()
        if fmt: ws[f'{col}{tr}'].number_format = fmt
    ws.row_dimensions[tr].height = 20
    ws.freeze_panes = 'A4'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
# INTERFACCIA UTENTE
# ══════════════════════════════════════════════════════════════════════════════

# Intestazione
st.markdown("""
<div class="intestazione">
    <h1>⚡ Ottimizzatore Consumi Energetici</h1>
    <p>Carica i prezzi GME · Imposta i vincoli · Ottimizza automaticamente i tuoi carichi</p>
</div>
""", unsafe_allow_html=True)

# ── Pannello laterale ─────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Parametri di configurazione")
    st.markdown("---")

    potenza_kw = st.number_input(
        "🔌 Potenza Impegnata (kW)",
        min_value=0.5, value=3.0, step=0.5,
        help="Potenza massima contrattuale. Casa: 3 kW · PMI: 15–50 kW · Industria: 100–5000 kW"
    )

    quota_flessibile = st.slider(
        "🔄 Quota Carico Flessibile (%)",
        min_value=10, max_value=100, value=70, step=5,
        help="Percentuale dei consumi che puoi spostare (es. lavatrice, lavastoviglie, pompa di calore)"
    )

    st.markdown("**🕐 Finestra oraria operativa**")
    st.caption("Ore in cui è possibile spostare i carichi (es. turno di lavoro)")
    col1, col2 = st.columns(2)
    with col1:
        ora_inizio = st.number_input("Dalle ore", min_value=1, max_value=24, value=1, step=1,
                                     help="Prima ora operativa (es. 8 per turno mattina)")
    with col2:
        ora_fine = st.number_input("Alle ore", min_value=1, max_value=24, value=24, step=1,
                                   help="Ultima ora operativa (es. 19 per fabbrica monturno)")

    giorni_anno = st.number_input(
        "📅 Giorni lavorativi / anno",
        min_value=1, max_value=365, value=365, step=1
    )

    st.markdown("---")
    st.markdown("### 📊 Consumi orari tipici")
    st.caption("Inserisci i kWh consumati ora per ora")

    consumi_default = [
        0.1, 0.1, 0.1, 0.1, 0.1, 0.1,
        0.5, 1.5, 1.2, 1.0, 0.8, 1.5,
        1.2, 0.8, 0.8, 1.0, 1.2, 2.0,
        2.5, 2.0, 1.5, 1.0, 0.5, 0.2
    ]

    if 'domanda' not in st.session_state:
        st.session_state.domanda = consumi_default.copy()

    with st.expander("✏️ Modifica consumi ora per ora"):
        nuova_domanda = []
        for h in range(24):
            val = st.number_input(
                f"Ora {h+1:02d}:00",
                min_value=0.0,
                value=float(st.session_state.domanda[h]),
                step=0.05, key=f"d_{h}"
            )
            nuova_domanda.append(val)
        st.session_state.domanda = nuova_domanda

# ── Area principale ───────────────────────────────────────────────────────────
col_carica, col_istruzioni = st.columns([3, 2])

with col_carica:
    st.markdown('<div class="titolo-sezione">📁 Carica file prezzi GME</div>', unsafe_allow_html=True)
    file_caricato = st.file_uploader(
        "Trascina qui il file Excel scaricato dal sito GME (formato MGP-PUNPUN)",
        type=['xlsx', 'xls'],
        label_visibility="visible"
    )

with col_istruzioni:
    st.markdown('<div class="titolo-sezione">💡 Come scaricare i prezzi da GME</div>', unsafe_allow_html=True)
    st.markdown("""
    <div class="riquadro-istruzioni">
        <p>1. Vai su <strong>mercatoelettrico.it</strong></p>
        <p>2. Scegli: Dati e Pubblicazioni → Download → <strong>MGP</strong></p>
        <p>3. Seleziona la data → scarica il file <strong>PUN</strong> in formato Excel</p>
        <p>4. Carica il file nel riquadro qui accanto</p>
    </div>
    """, unsafe_allow_html=True)

# ── Lettura prezzi ────────────────────────────────────────────────────────────
prezzi = None

if file_caricato:
    prezzi, errore = leggi_file_gme(file_caricato)
    if errore:
        st.error(f"❌ Errore nella lettura del file: {errore}")
        st.info("Assicurati che il file sia nel formato GME standard (colonne: Data, Ora, Periodo, €/MWh)")
        prezzi = None
    else:
        st.success(f"✅ File caricato correttamente — {len(prezzi)} ore di prezzi importate")

if prezzi is None:
    prezzi = [
        152.84, 150.0, 146.67, 145.96, 146.65, 146.49,
        165.12, 175.21, 165.14, 156.2,  147.48, 128.59,
        123.31, 128.23, 134.45, 143.95, 141.27, 164.79,
        185.65, 180.23, 159.56, 157.55, 145.55, 147.2
    ]
    if not file_caricato:
        st.info("ℹ️ Nessun file caricato — vengono usati i prezzi di esempio (05/03/2026). Carica il file GME per i dati reali.")

domanda = st.session_state.domanda

# ── Riepilogo prezzi del giorno ───────────────────────────────────────────────
st.markdown('<div class="titolo-sezione">📈 Prezzi del giorno</div>', unsafe_allow_html=True)

c1, c2, c3, c4, c5 = st.columns(5)
ora_min = int(np.argmin(prezzi)) + 1
ora_max = int(np.argmax(prezzi)) + 1
with c1: st.metric("Prezzo Medio", f"{np.mean(prezzi):.2f} €/MWh")
with c2: st.metric("🟢 Ora più Economica", f"Ore {ora_min:02d}:00", f"{min(prezzi):.2f} €/MWh")
with c3: st.metric("🔴 Ora più Costosa",   f"Ore {ora_max:02d}:00", f"{max(prezzi):.2f} €/MWh")
with c4: st.metric("Prezzo Minimo",  f"{min(prezzi):.2f} €/MWh")
with c5: st.metric("Prezzo Massimo", f"{max(prezzi):.2f} €/MWh")

df_prezzi = pd.DataFrame({
    'Ora': [f"{h:02d}:00" for h in range(1, 25)],
    'Prezzo (€/MWh)': prezzi
})
st.line_chart(df_prezzi.set_index('Ora'), color='#059669', height=200)

# ── Bottone ottimizzazione ────────────────────────────────────────────────────
st.markdown('<div class="titolo-sezione">🚀 Avvia ottimizzazione</div>', unsafe_allow_html=True)

col_btn, col_spazio = st.columns([2, 5])
with col_btn:
    avvia = st.button("▶  ESEGUI OTTIMIZZAZIONE", use_container_width=True)

if avvia or 'ottimizzato' in st.session_state:
    if avvia:
        with st.spinner("⏳ Calcolo in corso..."):
            ott, successo = esegui_ottimizzazione(
                domanda, prezzi,
                potenza_kw=potenza_kw,
                quota_flessibile=quota_flessibile / 100,
                ora_inizio=ora_inizio,
                ora_fine=ora_fine
            )
            st.session_state.ottimizzato = ott
            st.session_state.ott_successo = successo

    ott      = st.session_state.ottimizzato
    successo = st.session_state.ott_successo

    costo_attuale   = sum(d * p / 1000 for d, p in zip(domanda, prezzi))
    costo_ottimizzato = sum(o * p / 1000 for o, p in zip(ott, prezzi))
    risparmio       = costo_attuale - costo_ottimizzato
    pct_risparmio   = risparmio / costo_attuale * 100 if costo_attuale > 0 else 0

    if not successo:
        st.warning("⚠️ Il solver non ha trovato una soluzione ottimale con questi vincoli. Prova ad aumentare la potenza impegnata o la quota flessibile.")

    # Riquadri risultato
    st.markdown('<div class="titolo-sezione">💰 Risultati dell\'ottimizzazione</div>', unsafe_allow_html=True)
    r1, r2, r3 = st.columns(3)
    with r1:
        st.markdown(f"""
        <div class="card-risparmio">
            <div class="etichetta">💰 Risparmio Giornaliero</div>
            <div class="valore-grande">€ {risparmio:.4f}</div>
            <div class="sotto-testo">{pct_risparmio:.1f}% in meno rispetto al costo attuale</div>
        </div>""", unsafe_allow_html=True)
    with r2:
        st.markdown(f"""
        <div class="card-info">
            <div class="etichetta">📅 Risparmio Annuo Stimato</div>
            <div class="valore-grande">€ {risparmio * giorni_anno:.2f}</div>
            <div class="sotto-testo">calcolato su {giorni_anno} giorni/anno</div>
        </div>""", unsafe_allow_html=True)
    with r3:
        st.markdown(f"""
        <div class="card-info">
            <div class="etichetta">📊 Confronto Costi</div>
            <div class="valore-grande">€ {costo_ottimizzato:.4f}</div>
            <div class="sotto-testo">ottimizzato · vs € {costo_attuale:.4f} attuale</div>
        </div>""", unsafe_allow_html=True)

    # Tabella piano orario
    st.markdown('<div class="titolo-sezione">📋 Piano orario ottimizzato</div>', unsafe_allow_html=True)

    righe = []
    for i in range(24):
        h   = i + 1
        p   = prezzi[i]
        d   = domanda[i]
        o   = round(ott[i], 4)
        ca  = d * p / 1000
        co  = o * p / 1000
        delta = o - d
        freccia = "⬆️ aumenta" if delta > 0.01 else ("⬇️ riduci" if delta < -0.01 else "➡️ invariato")
        righe.append({
            "Ora":                    f"{h:02d}:00",
            "Prezzo (€/MWh)":         round(p, 2),
            "Consumo Attuale (kWh)":  round(d, 3),
            "Consumo Ottimizzato (kWh)": round(o, 3),
            "Azione":                 freccia,
            "Variazione (kWh)":       f"{delta:+.3f}",
            "Risparmio (€)":          f"€ {ca-co:+.5f}",
        })

    df_risultato = pd.DataFrame(righe)
    st.dataframe(df_risultato, use_container_width=True, height=540, hide_index=True)

    # Grafico confronto
    st.markdown('<div class="titolo-sezione">📊 Confronto consumi: attuale vs ottimizzato</div>', unsafe_allow_html=True)
    df_grafico = pd.DataFrame({
        'Ora': [f"{h:02d}:00" for h in range(1, 25)],
        'Consumo Attuale (kWh)':     [round(d, 3) for d in domanda],
        'Consumo Ottimizzato (kWh)': [round(o, 3) for o in ott],
    }).set_index('Ora')
    st.bar_chart(df_grafico, height=260, color=['#f87171', '#34d399'])

    # Matrice scenari
    st.markdown('<div class="titolo-sezione">🔬 Matrice scenari – risparmio giornaliero in €</div>', unsafe_allow_html=True)
    st.caption("Ogni cella mostra il risparmio ottenibile variando potenza impegnata e quota flessibile. Verde = conveniente, rosso = poco vantaggioso.")

    lista_flex = [50, 60, 70, 80, 90, 100]
    lista_pot  = [1.5, 3.0, 4.5, 6.0, 10.0]
    matrice    = {}
    for pot in lista_pot:
        riga = {}
        for fl in lista_flex:
            x, _ = esegui_ottimizzazione(domanda, prezzi, potenza_kw=pot, quota_flessibile=fl/100)
            c_ott = sum(xi * pi / 1000 for xi, pi in zip(x, prezzi))
            riga[f"Flex {fl}%"] = round(costo_attuale - c_ott, 4)
        matrice[f"{pot} kW"] = riga

    df_matrice = pd.DataFrame(matrice).T

    def colora_cella(val):
        """Colora le celle della matrice scenari senza matplotlib."""
        try:
            v = float(val)
        except:
            return ''
        if v >= 0.05:   return 'background-color: #6ee7b7; color: #064e3b; font-weight: 700'
        elif v >= 0.02: return 'background-color: #a7f3d0; color: #065f46; font-weight: 700'
        elif v >= 0.005: return 'background-color: #fef3c7; color: #92400e; font-weight: 700'
        else:           return 'background-color: #fecaca; color: #991b1b; font-weight: 700'

    st.dataframe(
        df_matrice.style
            .format("€ {:.4f}")
            .applymap(colora_cella),
        use_container_width=True
    )

    # Download report
    st.markdown("---")
    buf_excel = esporta_excel(
        prezzi, domanda, list(ott),
        {'potenza_kw': potenza_kw, 'quota_flessibile': quota_flessibile/100}
    )
    st.download_button(
        label="⬇️  SCARICA REPORT EXCEL",
        data=buf_excel,
        file_name="ottimizzazione_energia.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=False
    )

# Piè di pagina
st.markdown("---")
st.markdown(
    '<div class="pie-pagina">⚡ Ottimizzatore Consumi Energetici · Prezzi GME · Algoritmo LP (HiGHS solver)</div>',
    unsafe_allow_html=True
)
